from django.apps import apps
from django.contrib import messages
from django.db.models import Q
from django.forms import modelform_factory, modelformset_factory
from django.http import JsonResponse, HttpResponse
from django.shortcuts import redirect
from django.template.response import TemplateResponse
from django.template.defaultfilters import striptags
from django.utils.translation import ugettext
from django.views.generic import TemplateView, CreateView, UpdateView, View
from django.views.generic.base import TemplateResponseMixin, ContextMixin
from django.views.generic.edit import FormView
from django.urls import reverse

from json import loads
from openpyxl import Workbook
from openpyxl.utils import get_column_letter, units
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.writer.excel import save_virtual_workbook

from dal_select2.views import Select2QuerySetView

from GBEX_app.bulk_upload import bulk_import
from GBEX_app.forms import BulkUploadForm
from GBEX_app.helpers import field_to_string, get_free_id, model_to_list_list


class GBEXindex(TemplateView):
	template_name = 'GBEX_app/index.html'
	namespace = None

	def data_counts(self):
		all_numbers = {model.__name__: [model.objects.count(), 0] for model in apps.get_app_config('GBEX_app').get_models() if getattr(model, "model_kind", False) == 'GBEX_Page'}

		divider = max([x[0] for x in all_numbers.values()])
		if divider != 0:
			all_numbers = {key: [value[0], value[0] / divider * 100] for key, value in all_numbers.items()}
		return all_numbers


class GBEXCreateView(CreateView):
	model = None
	success_url = None
	form_class = None
	template_name = "GBEX_app/form.html"

	def form_valid(self, form):
		dups = form.cleaned_data['duplicater']
		form.instance.name = get_free_id(self.model)
		parent_inst = None
		if "parent_pk" in self.kwargs.keys():  # This is a batch model
			parent_inst = self.model.Parent.get_queryset().get(id=self.kwargs['parent_pk'])
			form.instance.Parent = parent_inst
			self.success_url = reverse(f'list_{self.model.__name__}', kwargs={'parent_pk': self.kwargs['parent_pk']})
		else:
			self.success_url = reverse(f'list_{self.model.__name__}')
		form.save()

		messages.success(request=self.request, message=f"Created new item: {form.instance.name}")
		for dup in range(dups - 1):
			form.instance.pk = None  # this makes the save mechanism save a new object
			#form.instance.organism_ptr = None  # this is needed for inherited models specifically organism inherited
			form.instance.name = get_free_id(self.model)
			if parent_inst:  # check if its a batch model
				form.instance.Parent = parent_inst
			form.save()
			messages.success(request=self.request, message=f"Created new item: {form.instance.name}")

		return super(GBEXCreateView, self).form_valid(form)


class GBEXUpdateView(UpdateView):
	"""
		UpdateView for a single field on a model
		It returns the new string value after successful post
	"""
	template_name = "GBEX_app/update.html"
	widgets = None

	def get_form_class(self):
		return modelform_factory(self.model, fields=[self.kwargs['column']], widgets=self.widgets)

	def form_valid(self, form):
		field_name = self.kwargs['column']
		form.save()
		model_instance = self.model.objects.get(pk=self.kwargs['pk'])
		return JsonResponse({"new_value": field_to_string(model_instance, field_name)})


class BulkUploadView(FormView):
	template_name = 'GBEX_app/bulk_upload.html'
	form_class = BulkUploadForm

	def form_valid(self, form):
		if not form.cleaned_data['file'].name.endswith("xlsx"):
			# we only support xlsx
			context = self.get_context_data()
			messages.warning(request=self.request, message=f"We only support <b>xlsx</b> files. You uploaded an {form.cleaned_data['file'].name.split('.')[-1]} file.")
			return self.render_to_response(context)
		bad_rows_file, total_bad_rows, return_messages, total_good_rows = bulk_import(self.request.FILES['file'], form.cleaned_data['upload_type'])
		for message in return_messages:
			messages.warning(request=self.request, message=message)
		if total_bad_rows != 0:
			messages.warning(request=self.request, message="You have bad rows. Please fix and reupload.")
			response = HttpResponse(bad_rows_file, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
			response['Content-Disposition'] = 'attachment; filename="bad_rows.xlsx"'
			return response
		elif total_good_rows == 0:
			context = self.get_context_data()
			messages.warning(request=self.request, message="You have 0 good rows.")
			return self.render_to_response(context)
		else:
			context = self.get_context_data()
			context['imported_rows'] = total_good_rows
			return TemplateResponse(self.request, 'GBEX_app/perfect_bulk.html', context)


class BulkUpdateView(TemplateView):
	template_name = "GBEX_app/bulk_update.html"
	model = None
	widgets = None

	def get(self, request, *args, **kwargs):
		context = self.get_context_data(**kwargs)
		context['forms'] = self.get_formset()(queryset=self.model.objects.filter(id__in=kwargs['rids'].split(",")))
		return self.render_to_response(context)

	def get_formset(self):
		return modelformset_factory(self.model, fields=[self.kwargs['column']], widgets=self.widgets, extra=0)

	def post(self, request, *args, **kwargs):
		rids = kwargs['rids'].split(",")
		formset = self.get_formset()
		infs = formset(request.POST, request.FILES, queryset=self.model.objects.filter(id__in=rids))
		if infs.is_valid():
			insts = infs.save()
			results = {x.id: field_to_string(x, self.kwargs['column']) for x in insts}
			return JsonResponse({"new_values": results})
		else:
			context = self.get_context_data(**kwargs)
			context['forms'] = infs
			return self.render_to_response(context)


class GBEXList(TemplateResponseMixin, ContextMixin, View):
	"""
	Convert an entire model into a string array
	"""
	template_name = "GBEX_app/list.html"
	model = None

	def get(self, request, *args, **kwargs):
		mnl = self.model.__name__
		context = self.get_context_data(**kwargs)
		context['model_name'] = mnl
		context['model_order'] = self.model.order

		datafilter = {'archived': False}
		url_kwargs = {}
		# parent_pk is part of the batch system and if found will only fetch a subset of the model with Parent=parent_pk
		context['return_url'] = reverse("GBEXindex")
		context['return_text'] = "GBEX"
		if 'parent_pk' in self.kwargs.keys():
			parent_pk = self.kwargs['parent_pk']
			context['return_url'] = reverse(f"list_{self.model.Parent.field.remote_field.model.__name__}")
			context['return_text'] = self.model.Parent.get_queryset().get(id=parent_pk).name
			datafilter['Parent'] = parent_pk
			url_kwargs = {'parent_pk': parent_pk}

		context['create_url'] = reverse(f"create_{mnl}", kwargs=url_kwargs)
		context['export_excel_url'] = reverse(f"export_{mnl}", kwargs=url_kwargs)
		context['data'] = model_to_list_list(self.model.objects.filter(**datafilter))

		context['table_settings'] = request.user.profile.table_settings
		context['settings_id'] = request.user.profile.id
		context['col_html_string'] = self.model.col_html_string
		context['col_read_only'] = self.model.col_read_only

		return self.render_to_response(context)


class GBEXAutocomplete(Select2QuerySetView):
	model = None
	search_fields = []
	limit_choices_to = {}

	def has_add_permission(self, request):
		return True

	def get_queryset(self):
		qs = self.model.objects.filter(**self.limit_choices_to)
		if self.q:
			filter_q = Q()
			for field in self.search_fields:
				filter_q |= Q(**{f'{field}__icontains': self.q})
			qs = qs.filter(filter_q)

		return qs

	def get_create_option(self, context, q):
		"""Form the correct create_option to append to results."""
		create_option = []
		display_create_option = False
		if self.create_field and q and not self.model.objects.filter(**{f"{self.create_field}__iexact": self.q}).exists():
			page_obj = context.get('page_obj', None)
			if page_obj is None or page_obj.number == 1:
				display_create_option = True

		if display_create_option and self.has_add_permission(self.request):
			create_option = [{
				'id': q,
				'text': ugettext('Create "%(new_value)s"') % {'new_value': q},
				'create_id': True,
			}]
		return create_option


class ExcelExportView(View):
	model = None

	def post(self, request, *args, **kwargs):
		rids = loads(request.body)['rids']
		if rids:
			# find all model instances with an id in rids
			objects = self.model.objects.filter(id__in=rids)
		else:
			# return all model instances from namespace
			datafilter = {'archived': False}
			# parent_pk is part of the batch system and if found will only fetch a subset of the model with Parent=parent_pk
			if 'parent_pk' in kwargs.keys():
				datafilter['Parent'] = self.kwargs['parent_pk']
			objects = self.model.objects.filter(**datafilter)

		# turn objects into a workbook
		data = model_to_list_list(objects)
		wb = Workbook()
		ws = wb.active
		ws.title = self.model.__name__

		ws.append(self.model.order)
		for row in data:
			ws.append([striptags(x) for x in row])  # We are removing all HTML tags...THIS MAY NOT ALWAYS BE DESIRED

		tab = Table(displayName="Table1", ref=f"A1:{get_column_letter(len(self.model.order))}{len(data)}")
		style = TableStyleInfo(
			name="TableStyleMedium9", showFirstColumn=True, showLastColumn=False,
			showRowStripes=True, showColumnStripes=False
		)
		tab.tableStyleInfo = style
		ws.add_table(tab)

		# set column width per user settings
		up = apps.get_model('GBEX_app', 'Profile').objects.filter(user=request.user)
		if up:
			json_data = loads(up[0].table_settings)
			if self.model.__name__ in json_data:
				if 'column_widths' in json_data[self.model.__name__]:
					cws = json_data[self.model.__name__]['column_widths']
					for i, col_name in enumerate(self.model.order):
						if col_name in cws:
							ws.column_dimensions[get_column_letter(i + 1)].width = units.pixels_to_points(cws[col_name]) / 6

		response = HttpResponse(
			save_virtual_workbook(wb),
			content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
		response['Content-Disposition'] = 'attachment; filename="export.xlsx"'
		return response


class ArchiveView(View):
	model = None

	def get(self, request, *args, **kwargs):
		if kwargs['rids']:
			# find all model instances with id in rids
			ids = kwargs['rids'].split(",")
			# set them all to archived
			self.model.objects.filter(id__in=ids).update(archived=True)
		# refresh page
		url_kwargs = {}
		if 'parent_pk' in kwargs.keys():
			url_kwargs['parent_pk'] = kwargs['parent_pk']

		return redirect(reverse(f"list_{self.model.__name__}", kwargs=url_kwargs))
