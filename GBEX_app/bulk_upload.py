from re import compile

from django.core.exceptions import ObjectDoesNotExist, ValidationError
from django.db.transaction import atomic
from django.db.utils import IntegrityError
from django.apps import apps

from openpyxl import load_workbook, Workbook
from openpyxl.writer.excel import save_virtual_workbook

from GBEX_app.helpers import return_stripped

bulk_importable_models = {model.__name__: model for model in apps.get_app_config('GBEX_app').get_models()}


def pre_import(wb, bad_book, return_messages, create):
	"""
	This function checks the import excel and discovers column info and prepares bad_book and if create mode then creates items
	:param wb: import excel sheet
	:param bad_book: return excel sheet with errors
	:param return_messages: list of messages to return to the user
	:param create: bool indicating whether we are in create or edit mode
	:return: good_sheets, return_messages, bad_book, bad_rows
	"""
	good_sheets = {}  # title: {'sheet': sheet, 'model': model, 'fkeys': list, 'm2ms': list, 'error_write_column': number, 'field_to_column_map': dict}
	bad_rows = {}  # name: error

	first_sheet = True
	for this_sheet in wb:
		sheet_title = this_sheet.title
		if sheet_title not in bulk_importable_models.keys():
			return_messages.append(f"Sheet title '{this_sheet.title}' doesnt match any database. Skipping.")
			continue

		if first_sheet:
			bad_sheet = bad_book.active
			bad_sheet.title = sheet_title
			first_sheet = False
		else:
			bad_sheet = bad_book.create_sheet(sheet_title)

		model = bulk_importable_models[sheet_title]
		fkeys = {}  # field_name_lower: rel_model
		m2ms = {}  # field_name_lower: rel_model
		required_fields = {}  # fields with blank=False
		model_order = [x for x in model.order if
					   x != 'id' and 'file' not in x.lower()]  # remove id...nobody gets to change that and we dont support files right now
		# figure out which fields are relational
		for field in model._meta.get_fields():
			if field.name in model_order:
				if field.is_relation:
					if field.many_to_many:
						m2ms[field.name] = [field.name, field.remote_field.model]
					else:
						fkeys[field.name] = field.remote_field.model
				if not field.blank:
					required_fields[field.name] = field.default  # store the default value


		# Go through columns
		# Check if names are valid model fields and keep going until you spot two empty headers, stop, and write errors in next column
		# could also just go with openpyxl sheet[row] but then I would be assuming that people make sane excel files
		field_to_column_map = {}
		empty_header = 0
		for col_idx in range(1, len(model_order)+99):  # nobody will EVER make more than 99 superflues columns...wink wink...dataloss
			cell = this_sheet.cell(column=col_idx, row=1)
			header = return_stripped(cell.value)
			bad_sheet.cell(column=col_idx, row=1).value = header
			if not header:
				empty_header += 1
				if empty_header == 2:
					error_write_column = cell.column_letter
					break
				else:
					continue

			if header in model_order:
				# ok its a valid field name
				field_to_column_map[header] = col_idx
			elif 'file' in header:
				return_messages.append(f"bulk upload for file columns are not implemented: {sheet_title}:{cell.coordinate}:'{header}'. Skipping.")
			elif header != 'id':
				# Return message about unrecognized column names
				return_messages.append(f"Unknown column header: {sheet_title}:{cell.coordinate}:'{header}'. Skipping.")
		created_insts = {}
		if create:
			row_number = 2  # row 2 is first row after header (openpyxl, like excel, uses 1 based rows)
			prefix = model.symbol
			good_name = compile(f"^{prefix}\\d+$")
			while True:
				name = return_stripped(this_sheet.cell(column=field_to_column_map['name'], row=row_number).value)
				# check for empty row
				if not name:
					break
				try:
					# check if name is ok for this project/model
					if not good_name.match(name):
						raise ValueError(f"Name '{name}' does not adhere to naming scheme: '{prefix}N', where N is a number")
					reqs = {}
					for reqqed, default in required_fields.items():
						value = return_stripped(
							this_sheet.cell(column=field_to_column_map[reqqed], row=row_number).value)
						if not value:
							value = default
						if reqqed in fkeys.keys():
							finst = fkeys[reqqed].objects.get(name=value)
							reqs[reqqed] = finst
						else:
							reqs[reqqed] = value
					reqs['name'] = name
					# create
					with atomic():
						# create the smallest object possible, aka name + blank=False fields
						created_insts[name] = model.objects.create(**reqs)
				except (IntegrityError, ObjectDoesNotExist, ValueError, AttributeError, NameError, ValidationError) as e:
					bad_rows[name] = e

				row_number += 1
		good_sheets[sheet_title] = [this_sheet, model, fkeys, m2ms, error_write_column, field_to_column_map, created_insts]
	return good_sheets, bad_rows


def bulk_import(excel_file, upload_type):
	wb = load_workbook(excel_file, data_only=True)
	bad_book = Workbook()  # create a excel workbook containing all failed rows
	return_messages = []
	total_bad_rows = 0
	total_good_rows = 0

	good_sheets, bad_rows = pre_import(wb, bad_book, return_messages, upload_type == 'create')

	for sheet_title, (this_sheet, model, fkeys, m2ms, error_write_column, field_to_column_map, created_insts) in good_sheets.items():
		if len(field_to_column_map.keys()) == 0:
			# we have no columns...
			return_messages.append(f"Sheet '{sheet_title}' has no valid columns. Skipping.")
			continue
		row_number = 2  # row 2 is first row after header (openpyxl, like excel, uses 1 based rows)
		bad_row_number = 2
		bad_sheet = bad_book[sheet_title]
		# then iterate over rows
		while True:
			name = return_stripped(this_sheet.cell(column=field_to_column_map['name'], row=row_number).value)
			# check for empty row
			if not name:
				total_bad_rows += bad_row_number - 2
				break

			try:
				with atomic():
					# check for bad row
					if name in bad_rows.keys():
						raise bad_rows[name]
					# update can only run on querysets...so we get a queryset

					if upload_type == 'create':
						instance_query = model.objects.filter(id=created_insts[name].id)
					else:
						instance_query = model.objects.filter(name=name)

					if instance_query.count() == 0:
						raise ValueError(f"Name '{name}' not found in database '{model.__name__}'")
					elif instance_query.count() != 1:
						raise ValueError("Name not unique")

					update_dict = {}
					m2m_updates = {}
					# go through columns
					for col_name, col_idx in field_to_column_map.items():
						# skip name columns. we are not updating that
						if col_name != 'name':
							value = return_stripped(
								this_sheet.cell(column=field_to_column_map[col_name], row=row_number).value)
							if value:
								# check if relational
								if col_name in fkeys.keys():
									# its foreign, write directly after finding it
									finst = fkeys[col_name].objects.get(name=value)
									update_dict[col_name] = finst
								elif col_name in m2ms.keys():
									# its lots of foreign, write in seperate step, just save for now
									finst_names = [x.strip() for x in value.split(",") if x.strip()]
									field_name = m2ms[col_name][0]
									rel_model = m2ms[col_name][1]
									m2m_updates[field_name] = []
									for finst_name in finst_names:
										# some fields are not project dependent..lets check
										m2m_updates[field_name].append(rel_model.objects.get(name=finst_name))
								else:
									# it is what it is
									update_dict[col_name] = value

					instance_query.update(**update_dict)
					inst = instance_query[0]
					for field_name, insts in m2m_updates.items():
						inst.__getattribute__(field_name).clear()  # remove any existing links
						inst.__getattribute__(field_name).add(*insts)  # add the new ones
					inst.full_clean()
					inst.save()
					total_good_rows += 1

			except (IntegrityError, ObjectDoesNotExist, ValidationError, ValueError, AttributeError, NameError, FileNotFoundError) as e:
				# copy entire row and write error info in "error_write_column"
				for cell in this_sheet[row_number]:
					pass
					bad_sheet[f"{cell.column_letter}{bad_row_number}"] = return_stripped(cell.value)
					bad_sheet[f"{cell.column_letter}{bad_row_number}"].hyperlink = cell.hyperlink

				error_type = e.__class__.__name__
				error_text = e.__str__()

				bad_sheet[f"{error_write_column}{bad_row_number}"] = f"{error_type}: {error_text}"
				bad_row_number += 1
				# delete row if we created it. Cant leave unfinished rows around
				if name in created_insts:
					created_insts[name].delete()

			row_number += 1

	return save_virtual_workbook(bad_book), total_bad_rows, return_messages, total_good_rows
